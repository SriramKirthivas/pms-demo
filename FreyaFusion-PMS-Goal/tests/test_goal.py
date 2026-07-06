import io

from openpyxl import Workbook, load_workbook

from conftest import ADMIN, EMPLOYEE, MANAGER, auth, token


def _sheet_bytes(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Goals"
    header = ["measure", "description", "baseCriteria", "pillar", "cadence", "defaultWeight", "employeeId"]
    ws.append(header)
    for r in rows:
        ws.append([r.get(h, "") for h in header])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_health(client):
    r = client.get("/api/health")
    assert r.status_code == 200 and r.json()["service"] == "pm-goal"


def test_framework_requires_admin(client):
    r = client.post("/api/pm-goal/framework", json={"fiscalYear": "FY26-27"}, headers=auth(token(*MANAGER)))
    assert r.status_code == 403 and r.json()["code"] == "FORBIDDEN"


def test_framework_weights_must_total_100(client):
    r = client.post("/api/pm-goal/framework",
                    json={"fiscalYear": "FY26-27", "teamWeightPct": 70, "individualWeightPct": 40},
                    headers=auth(token(*ADMIN)))
    assert r.status_code == 400 and r.json()["code"] == "PARAM_INVALID"


def test_framework_derives_periods(client):
    r = client.post("/api/pm-goal/framework",
                    json={"fiscalYear": "FY26-27", "activeCadences": ["QUARTERLY", "ANNUAL"],
                          "teamWeightPct": 60, "individualWeightPct": 40},
                    headers=auth(token(*ADMIN)))
    assert r.status_code == 200 and r.json()["code"] == 200
    codes = sorted(p["code"] for p in r.json()["data"]["periods"])
    assert codes == ["ANNUAL", "Q1", "Q2", "Q3", "Q4"]


def test_cascade_requires_section_weight_10(client):
    a = token(*ADMIN)
    g1 = client.post("/api/pm-goal/goals",
                     json={"fiscalYear": "FY26-27", "pillar": "TEAM_GOAL", "measure": "G1", "defaultWeight": 5},
                     headers=auth(a)).json()["data"]
    r = client.post(f"/api/pm-goal/goals/{g1['id']}/cascade", json={"employeeIds": ["David Chen"]}, headers=auth(a))
    assert r.status_code == 400 and r.json()["code"] == "PARAM_INVALID"
    client.post("/api/pm-goal/goals",
                json={"fiscalYear": "FY26-27", "pillar": "TEAM_GOAL", "measure": "G2", "defaultWeight": 5},
                headers=auth(a))
    r2 = client.post(f"/api/pm-goal/goals/{g1['id']}/cascade", json={"employeeIds": ["David Chen"]}, headers=auth(a))
    assert r2.status_code == 200 and r2.json()["data"]["assignments"][0]["status"] == "PENDING_ACCEPTANCE"
    assert len(r2.json()["data"]["created"]) == 1 and r2.json()["data"]["failed"] == []


def test_bilateral_acceptance_activates(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    client.post("/api/pm-goal/goals",
                json={"fiscalYear": "FY40", "pillar": "TEAM_GOAL", "measure": "A", "defaultWeight": 6}, headers=auth(a))
    g = client.post("/api/pm-goal/goals",
                    json={"fiscalYear": "FY40", "pillar": "TEAM_GOAL", "measure": "B", "defaultWeight": 4},
                    headers=auth(a)).json()["data"]
    aid = client.post(f"/api/pm-goal/goals/{g['id']}/cascade",
                      json={"employeeIds": ["David Chen"]}, headers=auth(a)).json()["data"]["created"][0]
    r1 = client.post(f"/api/pm-goal/assignments/{aid}/accept", headers=auth(e)).json()["data"]
    assert r1["status"] == "PENDING_ACCEPTANCE"
    r2 = client.post(f"/api/pm-goal/assignments/{aid}/accept", headers=auth(a)).json()["data"]
    assert r2["status"] == "ACTIVE" and r2["isActive"] is True
    actions = [t["action"] for t in client.get(f"/api/pm-goal/assignments/{aid}/audit", headers=auth(a)).json()["data"]]
    assert "CREATED" in actions and "CASCADED" in actions and actions.count("ACCEPTED") >= 2


def test_lock_period_locks_active_assignments(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    client.post("/api/pm-goal/framework",
                json={"fiscalYear": "FY50", "activeCadences": ["QUARTERLY"],
                      "teamWeightPct": 60, "individualWeightPct": 40}, headers=auth(a))
    client.post("/api/pm-goal/goals",
                json={"fiscalYear": "FY50", "pillar": "TEAM_GOAL", "measure": "A",
                      "defaultWeight": 6, "cadence": "QUARTERLY"}, headers=auth(a))
    g = client.post("/api/pm-goal/goals",
                    json={"fiscalYear": "FY50", "pillar": "TEAM_GOAL", "measure": "B",
                          "defaultWeight": 4, "cadence": "QUARTERLY"}, headers=auth(a)).json()["data"]
    aid = client.post(f"/api/pm-goal/goals/{g['id']}/cascade",
                      json={"employeeIds": ["David Chen"]}, headers=auth(a)).json()["data"]["created"][0]
    client.post(f"/api/pm-goal/assignments/{aid}/accept", headers=auth(e))
    client.post(f"/api/pm-goal/assignments/{aid}/accept", headers=auth(a))

    periods = client.get("/api/pm-goal/periods", params={"fiscalYear": "FY50"}, headers=auth(a)).json()["data"]
    q1 = next(p for p in periods if p["code"] == "Q1")
    r = client.post(f"/api/pm-goal/periods/{q1['id']}/lock", headers=auth(a))
    assert r.status_code == 200 and r.json()["data"]["lockedAssignments"] == 1
    # re-locking the same period conflicts
    assert client.post(f"/api/pm-goal/periods/{q1['id']}/lock", headers=auth(a)).status_code == 409
    # non-admin cannot lock
    q2 = next(p for p in periods if p["code"] == "Q2")
    assert client.post(f"/api/pm-goal/periods/{q2['id']}/lock", headers=auth(token(*MANAGER))).status_code == 403


def test_system_assignments_read(client):
    a = token(*ADMIN)
    g = client.post("/api/pm-goal/goals",
                    json={"fiscalYear": "FY51", "pillar": "TEAM_GOAL", "measure": "A", "defaultWeight": 10,
                          "cadence": "QUARTERLY"}, headers=auth(a)).json()["data"]
    client.post(f"/api/pm-goal/goals/{g['id']}/cascade", json={"employeeIds": ["David Chen"]}, headers=auth(a))
    rows = client.get("/api/pm-goal/system/assignments",
                      params={"employeeId": "David Chen", "fiscalYear": "FY51"}).json()["data"]
    assert len(rows) == 1 and rows[0]["pillar"] == "TEAM_GOAL" and rows[0]["weight"] == 10


def _cascade_active_assignment(client, fiscal_year, admin, employee, weight=10, cadence="QUARTERLY"):
    """Helper: create a framework + single fully-weighted goal, cascade it to
    David Chen, and bilaterally accept it so it becomes ACTIVE."""
    client.post("/api/pm-goal/framework",
                json={"fiscalYear": fiscal_year, "activeCadences": ["QUARTERLY"],
                      "teamWeightPct": 60, "individualWeightPct": 40}, headers=auth(admin))
    g = client.post("/api/pm-goal/goals",
                    json={"fiscalYear": fiscal_year, "pillar": "TEAM_GOAL", "measure": "A",
                          "defaultWeight": weight, "cadence": cadence}, headers=auth(admin)).json()["data"]
    aid = client.post(f"/api/pm-goal/goals/{g['id']}/cascade",
                      json={"employeeIds": ["David Chen"]}, headers=auth(admin)).json()["data"]["created"][0]
    client.post(f"/api/pm-goal/assignments/{aid}/accept", headers=auth(employee))
    client.post(f"/api/pm-goal/assignments/{aid}/accept", headers=auth(admin))
    return aid


# --- TC-G-16: Unlock approval chain ---

def test_unlock_request_and_approve_reopens_period(client):
    a, m, e = token(*ADMIN), token(*MANAGER), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY60", a, e)
    periods = client.get("/api/pm-goal/periods", params={"fiscalYear": "FY60"}, headers=auth(a)).json()["data"]
    q1 = next(p for p in periods if p["code"] == "Q1")
    client.post(f"/api/pm-goal/periods/{q1['id']}/lock", headers=auth(a))

    # manager requests unlock
    r = client.post(f"/api/pm-goal/periods/{q1['id']}/unlock-request",
                    json={"reason": "Late correction needed"}, headers=auth(m))
    assert r.status_code == 200 and r.json()["data"]["status"] == "PENDING"
    req_id = r.json()["data"]["id"]

    # the frontend must be able to discover the pending request's real id
    # (it has no other way to learn it before deciding)
    pending = client.get(f"/api/pm-goal/periods/{q1['id']}/unlock-requests",
                         params={"status": "PENDING"}, headers=auth(a)).json()["data"]
    assert len(pending) == 1 and pending[0]["id"] == req_id

    # employee (not manager/admin) cannot request unlock
    r_bad = client.post(f"/api/pm-goal/periods/{q1['id']}/unlock-request",
                        json={"reason": "x"}, headers=auth(e))
    assert r_bad.status_code == 403

    # non-admin cannot decide
    assert client.post(
        f"/api/pm-goal/periods/{q1['id']}/unlock-request/{req_id}/decision",
        json={"decision": "APPROVE"}, headers=auth(m),
    ).status_code == 403

    # admin approves -> period unlocked
    r2 = client.post(
        f"/api/pm-goal/periods/{q1['id']}/unlock-request/{req_id}/decision",
        json={"decision": "APPROVE", "reason": "Approved for correction"}, headers=auth(a),
    )
    assert r2.status_code == 200 and r2.json()["data"]["status"] == "APPROVED"
    period_after = next(
        p for p in client.get("/api/pm-goal/periods", params={"fiscalYear": "FY60"}, headers=auth(a)).json()["data"]
        if p["code"] == "Q1"
    )
    assert period_after["locked"] is False

    # approval restores locked assignments to ACTIVE (symmetric with lock),
    # otherwise corrected ratings still couldn't be submitted after unlock
    a_after = client.get(f"/api/pm-goal/assignments/{aid}", headers=auth(a)).json()["data"]
    assert a_after["status"] == "ACTIVE"

    # deciding twice conflicts
    assert client.post(
        f"/api/pm-goal/periods/{q1['id']}/unlock-request/{req_id}/decision",
        json={"decision": "APPROVE"}, headers=auth(a),
    ).status_code == 409

    actions = [t["action"] for t in client.get(f"/api/pm-goal/assignments/{aid}/audit", headers=auth(a)).json()["data"]]
    assert "UNLOCK_REQUESTED" in actions and "UNLOCKED" in actions


def test_unlock_request_reject_keeps_locked(client):
    a, m, e = token(*ADMIN), token(*MANAGER), token(*EMPLOYEE)
    _cascade_active_assignment(client, "FY61", a, e)
    periods = client.get("/api/pm-goal/periods", params={"fiscalYear": "FY61"}, headers=auth(a)).json()["data"]
    q1 = next(p for p in periods if p["code"] == "Q1")
    client.post(f"/api/pm-goal/periods/{q1['id']}/lock", headers=auth(a))
    req_id = client.post(f"/api/pm-goal/periods/{q1['id']}/unlock-request",
                         json={"reason": "oops"}, headers=auth(m)).json()["data"]["id"]

    r = client.post(f"/api/pm-goal/periods/{q1['id']}/unlock-request/{req_id}/decision",
                    json={"decision": "REJECT", "reason": "Not justified"}, headers=auth(a))
    assert r.status_code == 200 and r.json()["data"]["status"] == "REJECTED"
    period_after = next(
        p for p in client.get("/api/pm-goal/periods", params={"fiscalYear": "FY61"}, headers=auth(a)).json()["data"]
        if p["code"] == "Q1"
    )
    assert period_after["locked"] is True


# --- TC-G-17: Locked-period edit enforcement (409) ---

def test_locked_period_blocks_assignment_mutation(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY62", a, e)
    periods = client.get("/api/pm-goal/periods", params={"fiscalYear": "FY62"}, headers=auth(a)).json()["data"]
    q1 = next(p for p in periods if p["code"] == "Q1")
    client.post(f"/api/pm-goal/periods/{q1['id']}/lock", headers=auth(a))

    assert client.post(f"/api/pm-goal/assignments/{aid}/accept", headers=auth(e)).status_code == 409
    assert client.post(f"/api/pm-goal/assignments/{aid}/reject", headers=auth(e)).status_code == 409
    assert client.post(f"/api/pm-goal/assignments/{aid}/request-change",
                       json={"weight": 5}, headers=auth(e)).status_code == 409


# --- TC-G-18: Cascade validation response shape ---

def test_cascade_reports_created_and_failed(client):
    a = token(*ADMIN)
    g1 = client.post("/api/pm-goal/goals",
                     json={"fiscalYear": "FY63", "pillar": "TEAM_GOAL", "measure": "A", "defaultWeight": 10},
                     headers=auth(a)).json()["data"]
    r = client.post(f"/api/pm-goal/goals/{g1['id']}/cascade",
                    json={"employeeIds": ["David Chen", "", "David Chen", None]}, headers=auth(a))
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["created"] == [data["assignments"][0]["id"]]
    reasons = {f["employeeId"]: f["reason"] for f in data["failed"]}
    assert "" in reasons
    assert "David Chen" in [f["employeeId"] for f in data["failed"]]  # duplicate flagged


def test_cascade_owner_cannot_be_reviewer(client):
    a = token(*ADMIN)
    client.post("/api/pm-goal/goals",
                json={"fiscalYear": "FY64", "pillar": "TEAM_GOAL", "measure": "A", "defaultWeight": 10},
                headers=auth(a))
    g = client.get("/api/pm-goal/goals", params={"fiscalYear": "FY64", "pillar": "TEAM_GOAL"},
                   headers=auth(a)).json()["data"]["list"][0]
    r = client.post(f"/api/pm-goal/goals/{g['id']}/cascade",
                    json={"employeeIds": ["David Chen"], "reviewerIds": ["David Chen"]}, headers=auth(a))
    assert r.status_code == 409 and r.json()["code"] == "CONFLICT"


def test_cascade_multiple_reviewers(client):
    a = token(*ADMIN)
    client.post("/api/pm-goal/goals",
                json={"fiscalYear": "FY65", "pillar": "TEAM_GOAL", "measure": "A", "defaultWeight": 10},
                headers=auth(a))
    g = client.get("/api/pm-goal/goals", params={"fiscalYear": "FY65", "pillar": "TEAM_GOAL"},
                   headers=auth(a)).json()["data"]["list"][0]
    r = client.post(f"/api/pm-goal/goals/{g['id']}/cascade",
                    json={"employeeIds": ["David Chen"], "reviewerIds": ["Sarah Mitchell", "Nina Patel"]},
                    headers=auth(a))
    assert r.status_code == 200
    assert len(r.json()["data"]["created"]) == 1


# --- TC-G-19: Cross-service contracts (pm-eval RATED, status check, pm-score ACKNOWLEDGED) ---

def test_system_rated_event_appends_audit(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY66", a, e)
    r = client.post(f"/api/pm-goal/system/assignments/{aid}/rated",
                    json={"source": "SELF", "ratedBy": "David Chen", "evaluatedAt": "2026-07-01T00:00:00Z"})
    assert r.status_code == 200
    actions = [t["action"] for t in client.get(f"/api/pm-goal/assignments/{aid}/audit", headers=auth(a)).json()["data"]]
    assert "RATED" in actions

    assert client.post(
        "/api/pm-goal/system/assignments/does-not-exist/rated",
        json={"source": "SELF", "ratedBy": "David Chen", "evaluatedAt": "2026-07-01T00:00:00Z"},
    ).status_code == 404


def test_system_assignment_status_shape(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY67", a, e)
    r = client.get(f"/api/pm-goal/system/assignments/{aid}")
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["status"] == "ACTIVE" and data["ownerId"] == "David Chen"
    assert "reviewerIds" in data and "periodLocked" in data and data["periodLocked"] is False


def test_system_acknowledge_assignments(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY68", a, e)
    # fiscalYear arrives as an int (pm-score contract); this service stores an
    # opaque fiscal-year code string ("FY68") so acknowledge matches via
    # substring against the numeric year.
    r = client.post("/api/pm-goal/system/assignments/acknowledge",
                    json={"employeeId": "David Chen", "fiscalYear": 68})
    assert r.status_code == 200 and r.json()["data"]["updatedCount"] == 1
    assignment = client.get(f"/api/pm-goal/assignments/{aid}", headers=auth(a)).json()["data"]
    assert assignment["status"] == "ACKNOWLEDGED"

    # A year that doesn't appear in any fiscal-year code updates nothing.
    r2 = client.post("/api/pm-goal/system/assignments/acknowledge",
                     json={"employeeId": "David Chen", "fiscalYear": 999})
    assert r2.status_code == 200 and r2.json()["data"]["updatedCount"] == 0


# --- TC-G-25: Early completion (owner requests, manager approves/rejects) ---

def test_early_completion_approved(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY85", a, e)

    # Only the owner can request completion; must be ACTIVE.
    assert client.post(f"/api/pm-goal/assignments/{aid}/request-completion",
                       json={"note": "Delivered in month 1"}, headers=auth(a)).status_code == 403
    r = client.post(f"/api/pm-goal/assignments/{aid}/request-completion",
                    json={"note": "Delivered in month 1"}, headers=auth(e))
    assert r.status_code == 200 and r.json()["data"]["status"] == "COMPLETION_REQUESTED"

    # While pending, edits are blocked.
    assert client.post(f"/api/pm-goal/assignments/{aid}/request-change",
                       json={"weight": 10}, headers=auth(e)).status_code == 409

    # Manager approves -> COMPLETED.
    r2 = client.post(f"/api/pm-goal/assignments/{aid}/completion-decision",
                     json={"decision": "APPROVE", "note": "Confirmed"}, headers=auth(a))
    assert r2.status_code == 200 and r2.json()["data"]["status"] == "COMPLETED"

    actions = [t["action"] for t in client.get(f"/api/pm-goal/assignments/{aid}/audit", headers=auth(a)).json()["data"]]
    assert "COMPLETION_REQUESTED" in actions and "COMPLETED" in actions


def test_early_completion_rejected_returns_to_active(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY86", a, e)
    client.post(f"/api/pm-goal/assignments/{aid}/request-completion", json={}, headers=auth(e))
    r = client.post(f"/api/pm-goal/assignments/{aid}/completion-decision",
                    json={"decision": "REJECT", "note": "Not done yet"}, headers=auth(a))
    assert r.status_code == 200 and r.json()["data"]["status"] == "ACTIVE"
    # No pending request after a decision.
    assert client.post(f"/api/pm-goal/assignments/{aid}/completion-decision",
                       json={"decision": "APPROVE"}, headers=auth(a)).status_code == 409


# --- TC-G-20: Reassignment + lifecycle ---

def test_reassign_transfers_reviewer_role(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY69", a, e)
    r = client.post(f"/api/pm-goal/assignments/{aid}/reassign",
                    json={"newManagerId": "Sarah Mitchell"}, headers=auth(a))
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["reviewerId"] == "Sarah Mitchell"
    actions = [t["action"] for t in client.get(f"/api/pm-goal/assignments/{aid}/audit", headers=auth(a)).json()["data"]]
    assert "REASSIGNED" in actions


def test_leaver_lifecycle_closes_open_assignments(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY70", a, e)
    r = client.post("/api/pm-goal/participants/David Chen/lifecycle",
                    json={"eventType": "LEAVER", "fiscalYear": "FY70"}, headers=auth(a))
    assert r.status_code == 200
    assignment = client.get(f"/api/pm-goal/assignments/{aid}", headers=auth(a)).json()["data"]
    assert assignment["status"] == "CLOSED" and assignment["partialYear"] is True


def test_joiner_lifecycle_creates_partial_year_assignment(client):
    a = token(*ADMIN)
    client.post("/api/pm-goal/framework",
                json={"fiscalYear": "FY71", "activeCadences": ["QUARTERLY"],
                      "teamWeightPct": 60, "individualWeightPct": 40}, headers=auth(a))
    g = client.post("/api/pm-goal/goals",
                    json={"fiscalYear": "FY71", "pillar": "TEAM_GOAL", "measure": "A",
                          "defaultWeight": 10, "cadence": "QUARTERLY"}, headers=auth(a)).json()["data"]

    r = client.post("/api/pm-goal/participants/New Hire/lifecycle",
                    json={"eventType": "JOINER", "fiscalYear": "FY71", "goalIds": [g["id"]]}, headers=auth(a))
    assert r.status_code == 200
    data = r.json()["data"]
    assert len(data["assignments"]) == 1
    assert data["assignments"][0]["partialYear"] is True
    assert data["assignments"][0]["ownerId"] == "New Hire"


def test_lifecycle_requires_setter_role(client):
    e = token(*EMPLOYEE)
    r = client.post("/api/pm-goal/participants/David Chen/lifecycle",
                    json={"eventType": "LEAVER", "fiscalYear": "FY72"}, headers=auth(e))
    assert r.status_code == 403


# --- TC-G-21: Tenant onboarding (schema/seed only) ---

def test_tenant_onboarding_completes_synchronously(client):
    r = client.post("/api/pm-goal/tenant/onboarding")
    assert r.status_code == 200 and r.json()["data"]["status"] == "COMPLETED"
    r2 = client.post("/api/pm-goal/tenant/onboarding/query")
    assert r2.status_code == 200 and r2.json()["data"] == {"status": "COMPLETED"}


# --- TC-G-22: Goal listing pagination + assignment periodId filter ---

def test_goal_list_pagination(client):
    a = token(*ADMIN)
    for i in range(3):
        client.post("/api/pm-goal/goals",
                    json={"fiscalYear": "FY73", "pillar": "TEAM_GOAL", "measure": f"G{i}", "defaultWeight": 1},
                    headers=auth(a))
    r = client.get("/api/pm-goal/goals", params={"fiscalYear": "FY73", "pageNum": 1, "pageSize": 2},
                   headers=auth(a))
    data = r.json()["data"]
    assert data["total"] == 3 and len(data["list"]) == 2 and data["pageNum"] == 1


def test_assignments_filter_by_period_id(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    aid = _cascade_active_assignment(client, "FY74", a, e)
    assignment = client.get(f"/api/pm-goal/assignments/{aid}", headers=auth(a)).json()["data"]
    pid = assignment["periodId"]
    assert pid  # resolved a concrete period at cascade time

    matched = client.get("/api/pm-goal/assignments", params={"periodId": pid}, headers=auth(a)).json()["data"]
    assert any(x["id"] == aid for x in matched)
    unmatched = client.get("/api/pm-goal/assignments", params={"periodId": "nope"}, headers=auth(a)).json()["data"]
    assert all(x["id"] != aid for x in unmatched)


# --- TC-G-23: Goal sheet import / export ---

def test_import_goal_sheet_creates_goals_and_cascades(client):
    a = token(*ADMIN)
    content = _sheet_bytes([
        {"measure": "Ship feature X", "description": "Desc", "baseCriteria": "1/3/5",
         "pillar": "TEAM_GOAL", "cadence": "QUARTERLY", "defaultWeight": 6, "employeeId": "David Chen"},
        {"measure": "Reduce defects", "description": "Desc2", "baseCriteria": "1/3/5",
         "pillar": "TEAM_GOAL", "cadence": "QUARTERLY", "defaultWeight": 4, "employeeId": "David Chen"},
    ])
    files = {"file": ("goals.xlsx", content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post("/api/pm-goal/goals/import", params={"fiscalYear": "FY80"}, files=files, headers=auth(a))
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["goalsCreated"] == 2 and data["assignmentsCreated"] == 2

    rows = client.get("/api/pm-goal/assignments", params={"employeeId": "David Chen"}, headers=auth(a)).json()["data"]
    measures = {x["measure"] for x in rows if x["fiscalYear"] == "FY80"}
    assert measures == {"Ship feature X", "Reduce defects"}


def test_import_goal_sheet_is_idempotent(client):
    a = token(*ADMIN)
    content = _sheet_bytes([
        {"measure": "Ship feature X", "description": "Desc", "baseCriteria": "1/3/5",
         "pillar": "TEAM_GOAL", "cadence": "QUARTERLY", "defaultWeight": 10, "employeeId": "David Chen"},
    ])
    files = {"file": ("goals.xlsx", content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r1 = client.post("/api/pm-goal/goals/import", params={"fiscalYear": "FY81"}, files=files, headers=auth(a))
    assert r1.json()["data"]["assignmentsCreated"] == 1

    files2 = {"file": ("goals.xlsx", content,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = client.post("/api/pm-goal/goals/import", params={"fiscalYear": "FY81"}, files=files2, headers=auth(a))
    assert r2.json()["data"]["assignmentsCreated"] == 0
    assert r2.json()["data"]["skippedExisting"] == 1

    rows = client.get("/api/pm-goal/assignments", params={"employeeId": "David Chen"}, headers=auth(a)).json()["data"]
    assert len([x for x in rows if x["fiscalYear"] == "FY81"]) == 1


def test_import_rejects_section_not_totaling_10(client):
    a = token(*ADMIN)
    content = _sheet_bytes([
        {"measure": "Bad goal", "description": "", "baseCriteria": "",
         "pillar": "TEAM_GOAL", "cadence": "QUARTERLY", "defaultWeight": 5, "employeeId": "David Chen"},
    ])
    files = {"file": ("goals.xlsx", content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post("/api/pm-goal/goals/import", params={"fiscalYear": "FY82"}, files=files, headers=auth(a))
    assert r.status_code == 400 and r.json()["code"] == "PARAM_INVALID"
    assert "TEAM_GOAL" in r.json()["message"]

    # No partial records created for the rejected section.
    rows = client.get("/api/pm-goal/assignments", params={"employeeId": "David Chen"}, headers=auth(a)).json()["data"]
    assert len([x for x in rows if x["fiscalYear"] == "FY82"]) == 0


def test_export_goal_sheet_roundtrip(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    _cascade_active_assignment(client, "FY83", a, e, weight=10)
    r = client.get("/api/pm-goal/goals/export", params={"employeeId": "David Chen", "fiscalYear": "FY83"},
                   headers=auth(a))
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "measure"
    assert any(row[0] == "A" for row in rows[1:])


# --- TC-G-24: Employee directory (UAM stub) + access scoping ---

PRIYA = ("employee", "Priya Nair")


def test_tenant_onboarding_seeds_directory(client):
    r = client.post("/api/pm-goal/tenant/onboarding")
    assert r.status_code == 200
    assert r.json()["data"]["directorySeeded"] == 7
    # Re-onboarding is idempotent — no duplicate rows, no error.
    r2 = client.post("/api/pm-goal/tenant/onboarding")
    assert r2.json()["data"]["directorySeeded"] == 0

    people = client.get("/api/pm-goal/people", headers=auth(token(*ADMIN))).json()["data"]
    assert {p["id"] for p in people} == {
        "Nina Patel", "Sarah Mitchell", "Elena Ruiz", "David Chen", "Priya Nair", "Tom Baker", "Marcus Webb",
    }


def test_people_default_scope_is_my_team(client):
    client.post("/api/pm-goal/tenant/onboarding")
    mine = client.get("/api/pm-goal/people", headers=auth(token(*MANAGER))).json()["data"]
    ids = {p["id"] for p in mine}
    assert ids == {"Sarah Mitchell", "David Chen", "Priya Nair", "Tom Baker"}
    assert "Marcus Webb" not in ids and "Nina Patel" not in ids


def test_people_manager_cannot_view_another_managers_team(client):
    client.post("/api/pm-goal/tenant/onboarding")
    r = client.get("/api/pm-goal/people", params={"managerId": "Elena Ruiz"}, headers=auth(token(*MANAGER)))
    assert r.status_code == 403


def test_cascade_rejects_unknown_employee_once_directory_populated(client):
    client.post("/api/pm-goal/tenant/onboarding")
    a = token(*ADMIN)
    client.post("/api/pm-goal/goals",
                json={"fiscalYear": "FY90", "pillar": "TEAM_GOAL", "measure": "A", "defaultWeight": 10},
                headers=auth(a))
    g = client.get("/api/pm-goal/goals", params={"fiscalYear": "FY90", "pillar": "TEAM_GOAL"},
                   headers=auth(a)).json()["data"]["list"][0]
    r = client.post(f"/api/pm-goal/goals/{g['id']}/cascade",
                    json={"employeeIds": ["David Chen", "Ghost Employee"]}, headers=auth(a))
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["created"] and len(data["created"]) == 1
    failed_ids = {f["employeeId"]: f["reason"] for f in data["failed"]}
    assert "not found in directory" in failed_ids["Ghost Employee"]


def test_assignments_default_scope_hides_other_employees(client):
    a, e = token(*ADMIN), token(*EMPLOYEE)
    _cascade_active_assignment(client, "FY91", a, e)

    # David Chen (owner) sees his own assignment with no filter.
    mine = client.get("/api/pm-goal/assignments", headers=auth(e)).json()["data"]
    assert any(x["fiscalYear"] == "FY91" for x in mine)

    # An uninvolved employee sees nothing for FY91 with no filter.
    other = client.get("/api/pm-goal/assignments", headers=auth(token(*PRIYA))).json()["data"]
    assert not any(x["fiscalYear"] == "FY91" for x in other)

    # And can't pull David Chen's assignments via an explicit filter either.
    r = client.get("/api/pm-goal/assignments", params={"employeeId": "David Chen"}, headers=auth(token(*PRIYA)))
    assert r.status_code == 403
