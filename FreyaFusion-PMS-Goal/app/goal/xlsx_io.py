"""Goal sheet import / export (.xlsx).

Standard fiscal-year template with two sections — "Quarterly Team Goals" and
"Individual Annual Goals" — sharing the same column layout:

    measure | description | baseCriteria | pillar | cadence | defaultWeight | employeeId

Import groups rows by (employeeId, pillar, cadence), validates that each
section's weights total 10, creates/finds a DRAFT Goal per distinct
(measure, pillar, cadence, fiscalYear) key and cascades an assignment to the
row's employee. Re-importing an identical sheet is a no-op (idempotent on the
natural key (employeeId, measure, pillar, cadence, fiscalYear)).
"""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..common import internal
from ..common.auth import CurrentUser
from ..common.envelope import PARAM_INVALID, ApiError
from . import service
from .models import Goal, GoalAssignment, Participation

COLUMNS = ["measure", "description", "baseCriteria", "pillar", "cadence", "defaultWeight", "employeeId"]
SECTION_TOTAL = 10


def _read_rows(ws) -> list[dict]:
    """Read data rows from a worksheet whose first row is the header."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx = {name: header.index(name) for name in COLUMNS if name in header}
    missing = [c for c in COLUMNS if c not in col_idx]
    if missing:
        raise ApiError(400, PARAM_INVALID, f"Sheet '{ws.title}' is missing column(s): {', '.join(missing)}")

    out = []
    for raw in rows[1:]:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        rec = {name: raw[idx] for name, idx in col_idx.items()}
        if rec.get("measure") is None or str(rec["measure"]).strip() == "":
            continue
        out.append(rec)
    return out


def parse_workbook(content: bytes) -> list[dict]:
    """Parse every sheet in the workbook into a flat list of row dicts."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    rows: list[dict] = []
    for ws in wb.worksheets:
        rows.extend(_read_rows(ws))
    return rows


def _validate_sections(rows: list[dict], fiscal_year: str) -> dict[tuple, list[dict]]:
    """Group rows by (employeeId, pillar, cadence) and validate each
    section's weights total 10. Raises 400 naming the offending section."""
    sections: dict[tuple, list[dict]] = {}
    for r in rows:
        emp = str(r.get("employeeId") or "").strip()
        pillar = str(r.get("pillar") or "").strip()
        cadence = str(r.get("cadence") or "QUARTERLY").strip() or "QUARTERLY"
        if not emp:
            raise ApiError(400, PARAM_INVALID, "Row is missing employeeId")
        if not pillar:
            raise ApiError(400, PARAM_INVALID, f"Row for {emp} is missing pillar")
        key = (emp, pillar, cadence)
        sections.setdefault(key, []).append(r)

    for (emp, pillar, cadence), section_rows in sections.items():
        total = 0
        for r in section_rows:
            try:
                w = int(r.get("defaultWeight") or 0)
            except (TypeError, ValueError):
                raise ApiError(
                    400, PARAM_INVALID,
                    f"Section {pillar}/{cadence} for {emp} has a non-numeric defaultWeight",
                )
            total += w
        if total != SECTION_TOTAL:
            raise ApiError(
                400, PARAM_INVALID,
                f"Section {pillar}/{cadence} for employee {emp} (fiscalYear {fiscal_year}) "
                f"weights total {total}, must total {SECTION_TOTAL}",
            )
    return sections


def import_workbook(db: Session, content: bytes, fiscal_year: str, user: CurrentUser) -> dict:
    rows = parse_workbook(content)
    if not rows:
        raise ApiError(400, PARAM_INVALID, "Workbook contains no data rows")

    # Validate ALL sections before writing anything (no partial imports).
    sections = _validate_sections(rows, fiscal_year)

    created_goals = 0
    created_assignments = 0
    skipped_existing = 0
    failed: list[dict] = []

    for (emp, pillar, cadence), section_rows in sections.items():
        for r in section_rows:
            measure = str(r["measure"]).strip()
            description = str(r.get("description") or "").strip()
            base_criteria = str(r.get("baseCriteria") or "").strip()
            try:
                weight = int(r.get("defaultWeight") or 0)
            except (TypeError, ValueError):
                weight = 0

            # Idempotency key: (employeeId, measure, pillar, cadence, fiscalYear).
            existing_assignment = (
                db.query(GoalAssignment)
                .filter_by(
                    owner_id=emp, measure=measure, pillar=pillar,
                    cadence=cadence, fiscal_year=fiscal_year, is_delete=False,
                )
                .first()
            )
            if existing_assignment is not None:
                skipped_existing += 1
                continue

            goal = (
                db.query(Goal)
                .filter_by(
                    fiscal_year=fiscal_year, pillar=pillar, cadence=cadence,
                    measure=measure, is_delete=False,
                )
                .first()
            )
            if goal is None:
                goal = Goal(
                    fiscal_year=fiscal_year, pillar=pillar, cadence=cadence,
                    goal_type="OKR", measure=measure, description=description,
                    base_criteria=base_criteria, default_weight=max(0, min(10, weight)),
                    goal_status="DRAFT", setter_id=user.name, create_user=user.name,
                )
                db.add(goal)
                db.flush()
                created_goals += 1

            period = service._resolve_period(db, fiscal_year, cadence)
            a = GoalAssignment(
                goal_id=goal.id, fiscal_year=fiscal_year, period_id=period.id if period else "",
                owner_id=emp, setter_id=user.name, reviewer_id=user.name,
                pillar=pillar, cadence=cadence, goal_type=goal.goal_type,
                measure=measure, criteria=base_criteria, weight=max(0, min(10, weight)),
                competencies=[], assignment_status="PENDING_ACCEPTANCE",
                employee_acceptance="PENDING", manager_acceptance="PENDING",
                create_user=user.name,
            )
            db.add(a)
            db.flush()
            db.add(Participation(assignment_id=a.id, employee_id=emp, role="OWNER"))
            db.add(Participation(assignment_id=a.id, employee_id=user.name, role="SETTER"))
            db.add(Participation(assignment_id=a.id, employee_id=user.name, role="REVIEWER"))
            service._audit(db, a.id, user.name, "CREATED", f"Imported goal '{measure}' for {emp}")
            service._audit(db, a.id, user.name, "CASCADED", f"Imported cascade to {emp}")
            created_assignments += 1
            goal.goal_status = "CASCADED"

    db.commit()
    return {
        "fiscalYear": fiscal_year,
        "sections": len(sections),
        "goalsCreated": created_goals,
        "assignmentsCreated": created_assignments,
        "skippedExisting": skipped_existing,
        "failed": failed,
    }


def export_workbook(db: Session, employee_id: str, fiscal_year: str) -> bytes:
    rows = (
        db.query(GoalAssignment)
        .filter_by(owner_id=employee_id, fiscal_year=fiscal_year, is_delete=False)
        .order_by(GoalAssignment.pillar, GoalAssignment.cadence)
        .all()
    )

    # Source Goal metadata (description / base criteria) for each assignment.
    goal_ids = {a.goal_id for a in rows}
    goals_by_id = {}
    if goal_ids:
        for g in db.query(Goal).filter(Goal.id.in_(goal_ids)).all():
            goals_by_id[g.id] = g

    # Latest self/reviewer ratings + comments from pm-eval, keyed by
    # assignmentId. Best-effort — the sheet still exports (ratings blank) if
    # pm-eval is unreachable.
    ratings = internal.get_json(
        "eval", "/api/pm-eval/system/evaluations/latest", {"employeeId": employee_id},
    ) or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Goals"
    # Granular per-goal row: who owns it, who reviews it, where it stands, and
    # both sides' ratings + written views.
    header = [
        "Employee", "Manager", "Pillar", "Cadence", "Goal", "Description",
        "Measurement Criteria", "Weight", "Status", "Employee Accepted",
        "Manager Accepted", "Self Rating", "Employee's View",
        "Manager Rating", "Manager's View",
    ]
    ws.append(header)
    for a in rows:
        g = goals_by_id.get(a.goal_id)
        r = ratings.get(a.id) or {}
        self_r = r.get("self") or {}
        mgr_r = r.get("reviewer") or {}
        ws.append([
            a.owner_id,
            a.reviewer_id,
            a.pillar,
            a.cadence,
            a.measure,
            g.description if g else "",
            a.criteria or (g.base_criteria if g else ""),
            a.weight,
            a.assignment_status,
            a.employee_acceptance,
            a.manager_acceptance,
            self_r.get("rating", ""),
            self_r.get("comment", ""),
            mgr_r.get("rating", ""),
            mgr_r.get("comment", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
